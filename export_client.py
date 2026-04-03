"""
CLIENT DELIVERY PACKAGE — Generates a complete client-ready export
====================================================================

CAN BE USED TWO WAYS:

1. FROM COMMAND LINE:
   python export_client.py TestClient

2. FROM PYTHON (API server or other scripts):
   from export_client import run_export
   result = run_export("TestClient")
   # result = { "success": True, "delivery_folder": "...", "document_count": 5, "spreadsheet": "..." }

WHAT YOU GET:
- A timestamped delivery folder containing:
  - Subfolders per document type (Tenancy Agreements, Gas Safety, EICR, etc.)
  - PDFs renamed cleanly (e.g. "12 Oak Street - Tenancy Agreement.pdf")
  - An Excel index spreadsheet listing all documents with their verified fields

This is what you hand to the client.
"""

import sys
import json
import re
import shutil
from pathlib import Path
from datetime import datetime
import pdfplumber

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Fix:   pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────

BASE = Path(__file__).resolve().parent

DOC_TYPE_FOLDERS = {
    "Tenancy Agreement": "Tenancy Agreements",
    "Gas Safety Certificate": "Gas Safety Certificates",
    "EICR": "EICRs",
    "EPC": "EPCs",
    "General Document": "General Documents",
}


def get_category_folder_name(doc_type: str) -> str:
    """Map a document type to its export folder name."""
    return DOC_TYPE_FOLDERS.get(doc_type, f"{doc_type}s")


def sanitize_folder_name(name: str) -> str:
    """Sanitise a string for use as a Windows folder name."""
    if not name:
        return "Unsorted"
    for ch in '\\/:*?"<>|':
        name = name.replace(ch, "_")
    name = name.strip().rstrip(".")
    return name or "Unsorted"


def extract_pdf_text(pdf_path):
    """Extract all text from a PDF file using pdfplumber."""
    try:
        text = ""
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    except Exception as e:
        print(f"  Warning: Could not extract text from {pdf_path}: {e}")
        return ""


def extract_fields_from_text(doc_type: str, text: str) -> dict:
    """
    Parse document text and extract structured fields so the viewer can show
    key-value pairs (like the second picture) instead of only raw text.
    Verified fields from ReviewStation override these; this backfills empties.
    """
    if not text or not text.strip():
        return {}
    text = "\n" + text.replace("\r\n", "\n").replace("\r", "\n") + "\n"
    out = {}

    # Normalise to single spaces and preserve newlines for line-based patterns
    def line_capture(label_pattern: str, key: str) -> None:
        m = re.search(label_pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            val = m.group(1).strip()
            # Stop at next common section (e.g. next label or CLAUSE)
            val = re.split(r"\n\s*(?:CLAUSE|[\d]+\.\s+[A-Z]|$)", val, maxsplit=1)[0]
            val = " ".join(val.split())
            if val:
                out[key] = val

    if doc_type == "Tenancy Agreement":
        line_capture(r"(?:date\s+of\s+agreement|agreement(?:\s*/\s*signature)?\s*date)\s*:?\s*([^\n]+)", "agreement_date")
        line_capture(r"landlord\s*:?\s*([^\n]+)", "landlord_name")
        line_capture(r"tenant(?:s)?\s*\(?s\)?\s*:?\s*([^\n]+)", "tenant_full_name")
        line_capture(r"property\s*:?\s*([^\n]+)", "property_address")
        line_capture(r"term\s+start\s*:?\s*([^\n]+)", "start_date")
        line_capture(r"term\s+end\s*:?\s*([^\n]+)", "end_date")
        line_capture(r"(?:monthly\s+rent|rent\s+amount)\s*:?\s*([^\n]+)", "monthly_rent_amount")
        line_capture(r"deposit\s*:?\s*([^\n]+)", "deposit_amount")
        # Some docs use "Term: Start - End" on one line
        term_m = re.search(r"term\s*:?\s*([^\n]+?)\s+to\s+([^\n]+)", text, re.IGNORECASE)
        if term_m and not out.get("start_date"):
            out["start_date"] = " ".join(term_m.group(1).strip().split())
        if term_m and not out.get("end_date"):
            out["end_date"] = " ".join(term_m.group(2).strip().split())

    return out


def merge_fields(extracted: dict, verified: dict) -> dict:
    """Verified (from ReviewStation) overrides extracted; non-empty wins per key."""
    all_keys = set(extracted) | set(verified)
    merged = {}
    for k in all_keys:
        v = (verified.get(k) or "").strip()
        if v:
            merged[k] = v
        else:
            merged[k] = (extracted.get(k) or "").strip()
    return merged


# ──────────────────────────────────────────────
# COLLECT VERIFIED DOCUMENTS
# ──────────────────────────────────────────────

def collect_verified_docs(client_name: str):
    batches_path = BASE / "Clients" / client_name / "Batches"

    if not batches_path.exists():
        raise FileNotFoundError(f"Batches folder not found: {batches_path}")

    docs = []

    for date_folder in sorted(batches_path.iterdir()):
        if not date_folder.is_dir():
            continue

        for doc_folder in sorted(date_folder.iterdir()):
            if not doc_folder.is_dir():
                continue

            review_file = doc_folder / "review.json"
            if not review_file.exists():
                continue

            try:
                with review_file.open("r", encoding="utf-8") as f:
                    data = json.load(f)
            except (json.JSONDecodeError, Exception) as e:
                print(f"  WARNING: Could not read {review_file}: {e}")
                continue

            if data.get("status") != "Verified":
                print(f"  SKIP: {doc_folder.name} (status: {data.get('status', 'unknown')})")
                continue

            pdf_file = None
            for f in doc_folder.iterdir():
                if f.suffix.lower() == ".pdf":
                    pdf_file = f
                    break

            docs.append({
                "data": data,
                "doc_folder": doc_folder,
                "date_folder_name": date_folder.name,
                "pdf_file": pdf_file,
            })
            print(f"  ADDED: {doc_folder.name} ({data.get('doc_type', 'Unknown')})")

    return docs


# ──────────────────────────────────────────────
# GENERATE CLEAN PDF FILENAME
# ──────────────────────────────────────────────

def make_clean_filename(data: dict) -> str:
    fields = data.get("fields", {})
    doc_type = data.get("doc_type", "Document")
    doc_id = data.get("doc_id", "DOC-00000")

    # Try property address first
    name_part = fields.get("property_address", "").strip()

    # Fall back to document_title (General Documents)
    if not name_part:
        name_part = fields.get("document_title", "").strip()

    # Fall back to tenant name
    if not name_part:
        name_part = fields.get("tenant_full_name", "").strip()

    # Last resort: doc ID
    if not name_part:
        name_part = doc_id

    # Clean for filename
    for char in '/\\:*?"<>|':
        name_part = name_part.replace(char, "-")
    name_part = name_part.strip(". ")

    if len(name_part) > 80:
        name_part = name_part[:80].strip()

    return f"{name_part} - {doc_type}.pdf"


# ──────────────────────────────────────────────
# PACKAGE PDFs INTO DELIVERY FOLDER
# ──────────────────────────────────────────────

def mark_doc_exported(doc_folder: Path, exported_at: str):
    review_path = doc_folder / "review.json"
    if not review_path.exists():
        return
    try:
        with review_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        review_meta = data.get("review") or {}
        review_meta["exported_at"] = exported_at
        data["review"] = review_meta
        with review_path.open("w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"  WARNING: Could not mark exported_at for {doc_folder}: {e}")


def package_pdfs(docs: list, delivery_folder: Path):
    """
    Copy PDFs into the delivery folder organised by property -> document type.

    Returns:
        packaged: list of (doc_data, relative_path, batch_date) for spreadsheet
        archive_docs: list of dicts for archive_data.json
    """
    packaged = []
    archive_docs = []
    name_counts = {}

    for doc in docs:
        data = doc["data"]
        pdf_file = doc["pdf_file"]
        doc_type = data.get("doc_type", "Other")
        verified_fields = data.get("fields", {}) or {}

        if not pdf_file or not pdf_file.exists():
            print(f"  WARNING: No PDF found for {data.get('doc_id', '?')}, skipping")
            continue

        # Extract text once for full_text and for backfilling empty fields
        full_text = extract_pdf_text(pdf_file)
        extracted = extract_fields_from_text(doc_type, full_text)
        fields = merge_fields(extracted, verified_fields)

        # Determine property address and folder
        prop_address = (fields.get("property_address", "") or "Unsorted").strip()
        if not prop_address:
            prop_address = "Unsorted"
        property_folder = sanitize_folder_name(prop_address)

        category = get_category_folder_name(doc_type)

        property_dir = delivery_folder / property_folder
        subfolder = property_dir / category
        subfolder.mkdir(parents=True, exist_ok=True)

        clean_name = make_clean_filename({**data, "fields": fields})
        if clean_name in name_counts:
            name_counts[clean_name] += 1
            stem = clean_name.rsplit(".pdf", 1)[0]
            clean_name = f"{stem} ({name_counts[clean_name]}).pdf"
        else:
            name_counts[clean_name] = 1

        dest = subfolder / clean_name
        shutil.copy2(str(pdf_file), str(dest))

        # Mark export time in review.json (per document)
        exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            mark_doc_exported(doc["doc_folder"], exported_at)
            review_meta = data.get("review") or {}
            review_meta["exported_at"] = exported_at
            data["review"] = review_meta
        except Exception as e:
            print(f"  WARNING: Failed to update exported_at for {data.get('doc_id', '?')}: {e}")

        relative_path = dest.relative_to(delivery_folder).as_posix()
        data_with_merged = {**data, "fields": fields}
        packaged.append((data_with_merged, relative_path, doc["date_folder_name"]))
        print(f"  COPIED: {relative_path}")

        # Build archive doc entry (viewer JSON); fields = extracted + verified so viewer shows structure
        archive_docs.append({
            "doc_id": data.get("doc_id", ""),
            "doc_name": data.get("doc_name", ""),
            "doc_type": doc_type,
            "filename": clean_name,
            "pdf_path": relative_path,
            "fields": fields,
            "quality_score": data.get("quality_score", ""),
            "ocr_confidence": data.get("ocr_confidence", ""),
            "full_text": full_text,
            "property_address": prop_address,
        })

    return packaged, archive_docs


# ──────────────────────────────────────────────
# BUILD SPREADSHEET (DYNAMIC FIELDS)
# ──────────────────────────────────────────────

def build_spreadsheet(client_name: str, packaged: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Index"

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B3544")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill("solid", fgColor="F5F7FA")

    # Collect ALL unique field keys across all documents
    all_field_keys = []
    seen_keys = set()
    for doc_data, rel_path, batch_date in packaged:
        for key in doc_data.get("fields", {}).keys():
            if key not in seen_keys:
                all_field_keys.append(key)
                seen_keys.add(key)

    # Build columns: fixed + dynamic fields + review
    columns = [
        ("doc_id", "Doc ID"),
        ("doc_type", "Document Type"),
        ("batch_date", "Batch Date"),
        ("file_path", "File Location"),
    ]
    for key in all_field_keys:
        label = key.replace("_", " ").title()
        columns.append((f"field_{key}", label))
    columns.extend([
        ("reviewed_by", "Reviewed By"),
        ("reviewed_at", "Reviewed At"),
        ("exported_at", "Exported At"),
        ("review_notes", "Notes"),
    ])

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{client_name} — Verified Document Index"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1A1D27")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Documents: {len(packaged)}"
    sub_cell.font = Font(name="Arial", size=10, color="888888")
    sub_cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 22

    # Headers (row 4)
    header_row = 4
    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 28

    # Data rows
    for row_idx, (doc_data, rel_path, batch_date) in enumerate(packaged, start=header_row + 1):
        fields = doc_data.get("fields", {})
        review = doc_data.get("review", {})

        row_values = {
            "doc_id": doc_data.get("doc_id", ""),
            "doc_type": doc_data.get("doc_type", ""),
            "batch_date": batch_date,
            "file_path": rel_path,
            "reviewed_by": review.get("reviewed_by", ""),
            "reviewed_at": review.get("reviewed_at", ""),
            "exported_at": review.get("exported_at", ""),
            "review_notes": review.get("notes", ""),
        }
        for key in all_field_keys:
            row_values[f"field_{key}"] = fields.get(key, "")

        for col_idx, (key, label) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_values.get(key, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for col_idx, (key, label) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        if "address" in key:
            ws.column_dimensions[col_letter].width = 35
        elif "file_path" in key:
            ws.column_dimensions[col_letter].width = 45
        elif "notes" in key:
            ws.column_dimensions[col_letter].width = 30
        elif "name" in key:
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(packaged)}"

    return wb


def generate_archive_data(client_name: str, delivery_dir: Path, documents: list) -> Path:
    """
    Generate archive_data.json for the offline client viewer.

    documents: list of dicts with keys:
      - doc_id, doc_name, doc_type, filename, pdf_path (relative to delivery_dir)
      - fields (dict), quality_score, ocr_confidence, full_text, property_address
    """
    properties = {}

    for doc in documents:
        fields = doc.get("fields", {}) or {}
        prop_address = (doc.get("property_address") or fields.get("property_address", "") or "Unsorted").strip()
        if not prop_address:
            prop_address = "Unsorted"

        doc_type = doc.get("doc_type", "General Document")
        category = get_category_folder_name(doc_type)

        if prop_address not in properties:
            properties[prop_address] = {}
        if category not in properties[prop_address]:
            properties[prop_address][category] = []

        properties[prop_address][category].append({
            "doc_id": doc.get("doc_id", ""),
            "doc_name": doc.get("doc_name", ""),
            "doc_type": doc_type,
            "filename": doc.get("filename", ""),
            "pdf_path": doc.get("pdf_path", ""),
            "fields": fields,
            "full_text": doc.get("full_text", ""),
            "quality_score": doc.get("quality_score", ""),
            "ocr_confidence": doc.get("ocr_confidence", ""),
        })

    total_docs = sum(
        len(docs)
        for cats in properties.values()
        for docs in cats.values()
    )

    archive = {
        "client_name": client_name,
        "generated_date": datetime.now().strftime("%d %B %Y"),
        "version": "1.0",
        "total_documents": total_docs,
        "properties": properties,
    }

    output_path = delivery_dir / "archive_data.json"
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(archive, f, indent=2, ensure_ascii=False)

    print(f"  Generated archive_data.json ({total_docs} documents)")
    return output_path


def copy_viewer_assets(delivery_dir: Path) -> None:
    """Copy viewer.html (with archive data embedded) and instruction sheet into the delivery folder."""
    viewer_src = BASE / "viewer.html"
    archive_json = delivery_dir / "archive_data.json"
    if viewer_src.exists():
        try:
            html = viewer_src.read_text(encoding="utf-8")
            if archive_json.exists():
                data = json.loads(archive_json.read_text(encoding="utf-8"))
                # Embed so viewer works when opened as file:// (no fetch)
                payload = json.dumps(data, ensure_ascii=False).replace("</script>", "<\\/script>")
                placeholder = "window.ARCHIVE_DATA = null; /* Injected by export script when copying to Delivery folder */"
                if placeholder in html:
                    html = html.replace(placeholder, f"window.ARCHIVE_DATA = {payload}; /* embedded */", 1)
                else:
                    html = html.replace("window.ARCHIVE_DATA = null;", f"window.ARCHIVE_DATA = {payload};", 1)
            (delivery_dir / "viewer.html").write_text(html, encoding="utf-8")
            print("  Copied viewer.html (with archive data embedded)")
        except Exception as e:
            print(f"  WARNING: Failed to copy viewer.html: {e}")
    else:
        print("  Warning: viewer.html not found in system root")

    instructions_src = BASE / "instruction_sheet.pdf"
    if instructions_src.exists():
        try:
            shutil.copy2(str(instructions_src), str(delivery_dir / "Instructions.pdf"))
            print("  Copied Instructions.pdf")
        except Exception as e:
            print(f"  WARNING: Failed to copy Instructions.pdf: {e}")
    else:
        print("  Warning: instruction_sheet.pdf not found in system root")


# ──────────────────────────────────────────────
# MAIN EXPORT FUNCTION (callable from anywhere)
# ──────────────────────────────────────────────

def run_export(client_name: str) -> dict:
    """
    Run the full export for a client. Returns a result dict.
    
    Can be called from:
    - CLI (python export_client.py TestClient)
    - API server (from server.py)
    - Any other Python script
    
    Returns:
        {
            "success": True/False,
            "delivery_folder": "C:\\...\\Delivery_2026-02-23_1400",
            "document_count": 5,
            "spreadsheet": "TestClient_Document_Index_2026-02-23_1400.xlsx",
            "error": None or "error message"
        }
    """
    try:
        # Step 1: Collect verified documents
        docs = collect_verified_docs(client_name)

        if not docs:
            return {
                "success": False,
                "delivery_folder": None,
                "document_count": 0,
                "spreadsheet": None,
                "error": "No verified documents found. Mark documents as Verified in the Review Station first."
            }

        # Step 2: Create delivery folder
        exports_folder = BASE / "Clients" / client_name / "Exports"
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        delivery_folder = exports_folder / f"Delivery_{timestamp}"
        delivery_folder.mkdir(parents=True, exist_ok=True)

        # Step 3: Package PDFs (and collect archive data)
        packaged, archive_docs = package_pdfs(docs, delivery_folder)

        if not packaged:
            return {
                "success": False,
                "delivery_folder": str(delivery_folder),
                "document_count": 0,
                "spreadsheet": None,
                "error": "Documents were found but no PDFs could be packaged."
            }

        # Step 4: Generate spreadsheet
        wb = build_spreadsheet(client_name, packaged)
        spreadsheet_name = f"{client_name}_Document_Index_{timestamp}.xlsx"
        spreadsheet_path = delivery_folder / spreadsheet_name
        wb.save(str(spreadsheet_path))

        # Step 5: Generate archive_data.json for offline viewer
        generate_archive_data(client_name, delivery_folder, archive_docs)

        # Step 6: Copy viewer and instruction sheet (if present)
        copy_viewer_assets(delivery_folder)

        return {
            "success": True,
            "delivery_folder": str(delivery_folder),
            "document_count": len(packaged),
            "spreadsheet": spreadsheet_name,
            "error": None
        }

    except FileNotFoundError as e:
        return {
            "success": False,
            "delivery_folder": None,
            "document_count": 0,
            "spreadsheet": None,
            "error": str(e)
        }
    except Exception as e:
        return {
            "success": False,
            "delivery_folder": None,
            "document_count": 0,
            "spreadsheet": None,
            "error": f"Unexpected error: {e}"
        }


# ──────────────────────────────────────────────
# CLI ENTRY POINT (preserved — works exactly as before)
# ──────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python export_client.py <ClientName>")
        print("Example: python export_client.py TestClient")
        sys.exit(1)

    client_name = sys.argv[1]
    print("=" * 50)
    print("CLIENT DELIVERY PACKAGE")
    print(f"Client: {client_name}")
    print("=" * 50)
    print()

    result = run_export(client_name)

    if not result["success"]:
        print()
        print(f"FAILED: {result['error']}")
        sys.exit(1)

    # Summary
    print()
    print("=" * 50)
    print("DELIVERY PACKAGE READY")
    print("=" * 50)
    print(f"Location:    {result['delivery_folder']}")
    print(f"Documents:   {result['document_count']} PDFs")
    print(f"Spreadsheet: {result['spreadsheet']}")
    print()

    delivery_folder = Path(result['delivery_folder'])
    print("Folder contents:")
    for subfolder in sorted(delivery_folder.iterdir()):
        if subfolder.is_dir():
            count = len(list(subfolder.glob("*.pdf")))
            print(f"  /{subfolder.name}/ ({count} PDFs)")
    print(f"  {result['spreadsheet']}")
    print()
    print("Ready to deliver to client.")


if __name__ == "__main__":
    main()
